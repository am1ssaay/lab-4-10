import os
import openpyxl


class Integer:

    @classmethod
    def verify_value(cls, value):
        if type(value) != float and type(value) != int:
            raise TypeError(f"{value} Значение должно быть числом")
        if value < 0:
            raise TypeError(
                f"{value} Значение должно быть неотрицательным числом")

    def __set_name__(self, owner, name):
        self.name = "_" + name

    def __get__(self, instance, owner):
        return getattr(instance, self.name)

    def __set__(self, instance, value):
        self.verify_value(value)
        setattr(instance, self.name, value)


class Person:
    hours_norm = Integer()
    hours_norm = 168
    time = Integer()
    salary = Integer()

    def __init__(self, name: str, post: str, time: float, salary: float):
        self.verify_name(name)
        self.verify_post(post)

        self._name = name
        self._post = post
        self.time = time
        self.salary = salary

    def get_name(self):
        return self._name

    def get_post(self):
        return self._post

    def get_time(self):
        return self.time

    def get_salary(self):
        return self.salary

    def get_payment(self):
        return round(self.salary * (self.time / self.hours_norm), 2)

    def __str__(self) -> str:
        return f"ФИО: {self.get_name()} Должность: {self.get_post()} Отработанные часы: {self.time} Оклад: {self.salary} К выплате: {self.get_payment()}"

    def change_data(self, v, change):
        match v.split():
            case ["1"]:
                self.verify_name(change)
                self._name = change
            case ["2"]:
                self.verify_post(change)
                self._post = change
            case ["3"]:
                self.time = float(change)
            case ["4"]:
                self.salary = float(change)
            case _:
                print("Неверный ввод")
                return 1

    @classmethod
    def verify_name(cls, name: str):
        if type(name) != str:
            raise TypeError("ФИО должно быть строкой")

        f = name.split()
        if len(f) != 3:
            raise TypeError("Неверный формат ФИО")
        for s in f:
            if len(s) < 1:
                raise TypeError("В ФИО должен быть хотябы 1 символ")

    @classmethod
    def verify_post(cls, post: str):
        if type(post) != str:
            raise TypeError("Должность должна быть строкой")

        if len(post) < 1:
            raise TypeError("В названии должности должен быть хотябы 1 символ")

    @classmethod
    def change_hours_norm(cls, value):
        cls.hours_norm = float(value)


persons = []


def show_menu():
    print("1.Добавить сотрудника\n"
          "2.Удалить сотрудника из списка\n"
          "3.Вывести список сотрудников\n"
          "4.Изменить данные сотрудника\n"
          "5.Изменить норму часов\n"
          "6.Прочитать из файла\n"
          "7.Сохранить в файл\n"
          "8.Выйти")
    command = input("Введите номер опции: ")
    return command


def add_person():
    name_ = input("Введите ФИО сотрудника: ")
    post_ = input("Введите должность сотрудника: ")
    time_ = float(input("Введите отработанное время в часах сотрудника: "))
    salary_ = float(input("Введите оклад сотрудника: "))
    persons.append(Person(name_, post_, time_, salary_))
    return 1


def del_person():
    show_persons()
    p_id = input("Введите номер сотрудника: ")
    del persons[int(p_id)]
    return 1


def edit_person():
    show_persons()
    pers = int(
        input(
            "Введите номер сотрудника, у которого хотите отредактировать данные: "
        ))
    v = input("1.ФИО\n"
              "2.Должность\n"
              "3.Отработанное время\n"
              "4.Оклад\n"
              "Выберите, что хотите отредактировать:\n")
    change = input("Введите новые данные: ")
    persons[pers].change_data(v, change)
    return 1


def show_persons():
    if len(persons) < 1:
        print("Список сотрудников пуст")
        return 0
    for i in range(0, len(persons)):
        print(f"{i}. {persons[i]}")
    return 1


def read_from_file():
    path = input("Введите путь к файлу xlsx: ")
    if (not os.path.exists(path)):
        print("Такого файла не существует")
        return 0
    book = openpyxl.open(path, read_only=True)
    sheet = book.active
    for row in range(2, sheet.max_row + 1):
        name = sheet[row][0].value
        post = sheet[row][1].value
        time = float(sheet[row][2].value)
        salary = float(sheet[row][3].value)
        persons.append(Person(name, post, time, salary))
    book.close()
    return 1


def save_to_file():
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row=1, column=1).value = "ФИО"
    sheet.cell(row=1, column=2).value = "Должность"
    sheet.cell(row=1, column=3).value = "Отработанное время"
    sheet.cell(row=1, column=4).value = "Оклад"
    sheet.cell(row=1, column=5).value = "К выплате"

    for i in range(0, len(persons)):
        sheet.cell(row=i + 2, column=1).value = persons[i].get_name()
        sheet.cell(row=i + 2, column=2).value = persons[i].get_post()
        sheet.cell(row=i + 2, column=3).value = persons[i].get_time()
        sheet.cell(row=i + 2, column=4).value = persons[i].get_salary()
        sheet.cell(row=i + 2, column=5).value = persons[i].get_payment()
    book.save('data.xlsx')
    book.close()
    return 1

if __name__ == "__main__":
    while (True):
        command = show_menu()
        match command.split():
            case ["1"]:
                add_person()
            case ["2"]:
                del_person()
            case ["3"]:
                show_persons()
            case ["4"]:
                edit_person()
            case ["5"]:
                print(f"Текущая норма часов: {Person.hours_norm}")
                value = input("Введите новое значение нормы часов: ")
                Person.change_hours_norm(value)
            case ["6"]:
                read_from_file()
            case ["7"]:
                save_to_file()
            case ["8"]:
                os.system("cls")
                exit()
            case _:
                print("Неверный ввод")
                continue
